"""
Evaluation Module
=================
Evaluates predicted reaction graphs against ground-truth annotations.

Three matching modes:
* Hard  - all role members (reactants, conditions, products) must match.
* Soft  - only molecule-type (structure) members are compared.
* Hybrid - molecules must match exactly; text compared with edit distance.

Usage
-----
    python rxncaption/evaluate.py \
        --ground_truth_file data/ground_truth.json \
        --pred_file         data/prediction.json \
        --image_base_path   data/images \
        --output_dir        results/ \
        --mode all \
        --limit_vis 10
"""
from __future__ import annotations
import argparse, csv, json, logging, os
from collections import defaultdict
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import cv2
import numpy as np

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as patches
    HAS_MPL = True
except ImportError:
    HAS_MPL = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

def setup_logging(output_dir: str) -> str:
    log_path = os.path.join(output_dir, f"eval_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s",
                        handlers=[logging.FileHandler(log_path), logging.StreamHandler()])
    return log_path

# --- Edit distance ---
def _edit_distance(s1: str, s2: str) -> int:
    m, n = len(s1), len(s2)
    dp = list(range(n + 1))
    for i in range(1, m + 1):
        prev, dp[0] = dp[0], i
        for j in range(1, n + 1):
            temp = dp[j]
            if s1[i - 1] == s2[j - 1]:
                dp[j] = prev
            else:
                dp[j] = 1 + min(prev, dp[j], dp[j - 1])
            prev = temp
    return dp[n]

# --- IoU ---
def _iou(a, b):
    ax1, ay1, aw, ah = a; bx1, by1, bw, bh = b
    ax2, ay2 = ax1 + aw, ay1 + ah; bx2, by2 = bx1 + bw, by1 + bh
    ix1, iy1 = max(ax1, bx1), max(ay1, by1)
    ix2, iy2 = min(ax2, bx2), min(ay2, by2)
    inter = max(0, ix2 - ix1) * max(0, iy2 - iy1)
    union = aw * ah + bw * bh - inter
    return inter / union if union > 0 else 0.0

def get_bboxes_match(gt_boxes, pred_boxes, iou_threshold=0.5):
    n, m = len(gt_boxes), len(pred_boxes)
    if n == 0 or m == 0:
        return [-1] * n, m, np.zeros((n, m))
    iou_mat = np.array([[_iou(g, p) for p in pred_boxes] for g in gt_boxes])
    match = [-1] * n; used_pred = set()
    for gi in range(n):
        best_iou, best_pi = 0.0, -1
        for pi in range(m):
            if pi in used_pred: continue
            if iou_mat[gi, pi] > best_iou:
                best_iou, best_pi = iou_mat[gi, pi], pi
        if best_iou >= iou_threshold:
            match[gi] = best_pi; used_pred.add(best_pi)
    return match, m - len(used_pred), iou_mat

# --- Data model ---
class BBox:
    def __init__(self, record):
        self.id = record["id"]; self.category_id = record.get("category_id", 1)
        self.bbox = record.get("bbox", [0, 0, 0, 0]); self.text = record.get("text", "")
    @property
    def is_molecule(self): return self.category_id == 1
    @property
    def is_text(self): return self.category_id in (2, 3)

class Reaction:
    def __init__(self, rxn_dict, bbox_lookup):
        self.reactants = [bbox_lookup[i] for i in rxn_dict.get("reactants", []) if i in bbox_lookup]
        self.conditions = [bbox_lookup[i] for i in rxn_dict.get("conditions", []) if i in bbox_lookup]
        self.products = [bbox_lookup[i] for i in rxn_dict.get("products", []) if i in bbox_lookup]
    def all_members(self, mol_only=False):
        members = self.reactants + self.conditions + self.products
        return [b for b in members if b.is_molecule] if mol_only else members
    def molecule_bboxes(self): return [b.bbox for b in self.all_members(mol_only=True)]
    @property
    def bboxes(self): return [b.bbox for b in self.all_members()]
    def is_empty(self, mol_only=False): return len(self.all_members(mol_only=mol_only)) == 0
    def get_role(self, bbox_obj):
        if bbox_obj in self.reactants: return "reactants"
        if bbox_obj in self.conditions: return "conditions"
        if bbox_obj in self.products: return "products"
        return "unknown"

class ReactionImageData:
    def __init__(self, gold, pred, image_base_path, filter_recover=False):
        self.file_name = gold["file_name"]
        self.image_path = os.path.join(image_base_path, self.file_name)
        self.diagram_type = gold.get("diagram_type", "unknown")
        def _build(d): return {b["id"]: BBox(b) for b in d.get("bboxes", [])}
        gl, pl = _build(gold), _build(pred)
        self.gold_reactions = [Reaction(r, gl) for r in gold.get("reactions", [])]
        self.pred_reactions = [Reaction(r, pl) for r in pred.get("reactions", [])]

    def evaluate(self, mol_only=False, match_type="hard"):
        gt = [r for r in self.gold_reactions if not r.is_empty(mol_only)]
        pr = [r for r in self.pred_reactions if not r.is_empty(mol_only)]
        if not gt and not pr:
            return {"ignored": True, "tp": 0, "fp": 0, "fn": 0, "gold_total": 0, "pred_total": 0, "matches": []}
        matched_gt, matched_pr, matches = set(), set(), []
        for gi, gr in enumerate(gt):
            for pi, prr in enumerate(pr):
                if pi in matched_pr: continue
                if self._match(gr, prr, mol_only, match_type):
                    matched_gt.add(gi); matched_pr.add(pi); matches.append((gi, pi)); break
        tp = len(matches)
        return {"ignored": False, "tp": tp, "fp": len(pr) - tp, "fn": len(gt) - tp,
                "gold_total": len(gt), "pred_total": len(pr), "matches": matches}

    def _match(self, gr, pr, mol_only, match_type):
        if match_type == "hard": return self._hard(gr, pr)
        if match_type == "soft": return self._mol(gr, pr)
        if match_type == "hybrid": return self._hybrid(gr, pr)
        return False

    def _hard(self, gr, pr):
        for role in ("reactants", "conditions", "products"):
            gb = [b.bbox for b in getattr(gr, role)]; pb = [b.bbox for b in getattr(pr, role)]
            if len(gb) != len(pb): return False
            if gb:
                m, _, _ = get_bboxes_match(gb, pb)
                if any(v == -1 for v in m): return False
        return True

    def _mol(self, gr, pr):
        gb, pb = gr.molecule_bboxes(), pr.molecule_bboxes()
        if len(gb) != len(pb): return False
        if not gb: return True
        m, _, _ = get_bboxes_match(gb, pb)
        return all(v != -1 for v in m)

    def _hybrid(self, gr, pr, ratio=0.2):
        if not self._mol(gr, pr): return False
        for role in ("reactants", "conditions", "products"):
            gt_t = sorted(b.text for b in getattr(gr, role) if b.is_text)
            pr_t = sorted(b.text for b in getattr(pr, role) if b.is_text)
            if len(gt_t) != len(pr_t): return False
            used = set()
            for g in gt_t:
                ok = False
                for j, p in enumerate(pr_t):
                    if j in used: continue
                    d = _edit_distance(g, p); ml = max(len(g), len(p), 1)
                    if d / ml <= ratio: used.add(j); ok = True; break
                if not ok: return False
        return True

def compute_metrics(tp, gold, pred):
    p = tp / pred if pred > 0 else 0.0; r = tp / gold if gold > 0 else 0.0
    f1 = 2 * p * r / (p + r) if (p + r) > 0 else 0.0
    return {"precision": p, "recall": r, "f1": f1}

# --- Overall evaluation ---
def run_overall_evaluation(gold_images, pred_images, image_base_path, filter_recover=False):
    print("\n" + "=" * 60 + " Overall Evaluation " + "=" * 60)
    modes = {"Hard": {"mol_only": False, "match_type": "hard"},
             "Soft": {"mol_only": True, "match_type": "soft"},
             "Hybrid": {"mol_only": False, "match_type": "hybrid"}}
    gbn = {g["file_name"]: g for g in gold_images}
    pbn = {p["file_name"]: p for p in pred_images}
    for name, params in modes.items():
        tp = fp = fn = gt_total = pd_total = 0
        for k in sorted(set(gbn) & set(pbn)):
            data = ReactionImageData(gbn[k], pbn[k], image_base_path, filter_recover)
            res = data.evaluate(**params)
            if res.get("ignored"): continue
            tp += res["tp"]; fp += res["fp"]; fn += res["fn"]
            gt_total += res["gold_total"]; pd_total += res["pred_total"]
        m = compute_metrics(tp, gt_total, pd_total)
        print(f"  {name:<7} TP={tp:4d} FP={fp:4d} FN={fn:4d}  P={m['precision']*100:.2f}% R={m['recall']*100:.2f}% F1={m['f1']*100:.2f}%")
    print("=" * 140 + "\n")

# --- Visualization ---
ROLE_COLORS = {"reactants": "green", "conditions": "orange", "products": "blue", "unknown": "gray"}

def visualize_reaction_pair(img_rgb, pair, output_dir, file_name, item_idx, pair_idx, mol_only, match_type):
    if not HAS_MPL:
        print("[WARN] matplotlib not installed, skipping visualization."); return None
    nrows = 1 if match_type == "soft" else 2
    fig, axes = plt.subplots(nrows, 2, figsize=(16, 6 * nrows))
    if nrows == 1: axes = axes.reshape(1, 2)

    # Top-left: GT
    ax_gt = axes[0, 0]; ax_gt.imshow(img_rgb); ax_gt.set_axis_off()
    ax_gt.set_title("Ground Truth", fontsize=12)
    if pair["gt"]:
        for bb_obj in pair["gt"].all_members(mol_only):
            x, y, w, h = bb_obj.bbox
            role = pair["gt"].get_role(bb_obj)
            ax_gt.add_patch(patches.Rectangle((x, y), w, h, linewidth=2,
                            edgecolor=ROLE_COLORS.get(role, "gray"), facecolor="none"))

    # Top-right: Pred
    ax_pr = axes[0, 1]; ax_pr.imshow(img_rgb); ax_pr.set_axis_off()
    tc = "green" if "TP" in pair["status"] else ("red" if "FP" in pair["status"] else "gray")
    ax_pr.set_title("Prediction - " + pair["status"], fontsize=12, color=tc)
    if pair["pred"]:
        # Use the same member filter (mol_only) for both GT and Pred bbox lists
        # so that iou_mat column indices align with the enumeration below
        all_gt_bb = [b.bbox for b in pair["gt"].all_members(mol_only)] if pair["gt"] else []
        all_pr_bb = [b.bbox for b in pair["pred"].all_members(mol_only)]
        _, _, iou_mat = get_bboxes_match(all_gt_bb, all_pr_bb) if all_gt_bb else ([], 0, np.zeros((0, len(all_pr_bb))))
        for pid, bb_obj in enumerate(pair["pred"].all_members(mol_only)):
            x, y, w, h = bb_obj.bbox
            role = pair["pred"].get_role(bb_obj)
            max_iou = np.max(iou_mat[:, pid]) if iou_mat.size and pid < iou_mat.shape[1] else 0.0
            ls = "-" if max_iou >= 0.5 else "--"
            ax_pr.add_patch(patches.Rectangle((x, y), w, h, linewidth=2,
                            edgecolor=ROLE_COLORS.get(role, "gray"), linestyle=ls, facecolor="none"))
            ax_pr.text(x, y - 5, f"IoU:{max_iou:.2f}", color="darkgreen" if max_iou >= 0.5 else "darkred", fontsize=7)

    # Bottom: text comparison (hard/hybrid only)
    if nrows == 2:
        for ax in axes[1]: ax.axis("off")
        def role_texts(rxn):
            out = {"reactants": [], "conditions": [], "products": []}
            if not rxn: return out
            for role in out:
                for b in getattr(rxn, role):
                    if b.is_text and b.text: out[role].append(b.text.strip())
            return out
        def draw_txt(ax, td, label):
            lines = []
            for r in ("reactants", "conditions", "products"):
                if td[r]: lines.append(f"{r[:3].upper()}: " + ", ".join(td[r]))
            ax.text(0, 0.98, "\n".join(lines) if lines else "(none)", va="top", ha="left", fontsize=10, wrap=True)
            ax.set_title(label, fontsize=11, pad=2)
        draw_txt(axes[1, 0], role_texts(pair["gt"]), "GT Texts")
        draw_txt(axes[1, 1], role_texts(pair["pred"]), "Pred Texts")

    mode_label = {"soft": "Soft", "hybrid": "Hybrid", "hard": "Hard"}.get(match_type, "Hard")
    fig.suptitle(f"{file_name} - Pair {pair_idx+1} ({mode_label})", fontsize=14)
    plt.tight_layout(rect=[0, 0, 1, 0.94])

    vis_dir = os.path.join(output_dir, "visualizations")
    os.makedirs(vis_dir, exist_ok=True)
    safe = file_name.replace("/", "_").replace("\\", "_")
    out_path = os.path.join(vis_dir, f"{item_idx:04d}_{pair_idx:02d}_{mode_label.lower()}_{safe}.png")
    plt.savefig(out_path, dpi=110); plt.close(fig)
    return out_path

# --- Markdown visualization report ---
def generate_visualization_report(gold_images, pred_images, image_base_path, output_dir,
                                   vis_range=None, limit_vis=-1, mol_only=False,
                                   match_type="hard", filter_recover=False):
    if not HAS_MPL:
        print("[WARN] matplotlib not available. Skipping visualization."); return
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_dir = os.path.join(output_dir, f"md_{match_type}_{ts}")
    os.makedirs(report_dir, exist_ok=True)

    gbn = {g["file_name"]: g for g in gold_images}
    pbn = {p["file_name"]: p for p in pred_images}
    results = []
    for idx, fn in enumerate(sorted(set(gbn) & set(pbn))):
        data = ReactionImageData(gbn[fn], pbn[fn], image_base_path, filter_recover)
        res = data.evaluate(mol_only=mol_only, match_type=match_type)
        hard_res = data.evaluate(mol_only=False, match_type="hard")
        soft_res = data.evaluate(mol_only=True, match_type="soft")
        hybrid_res = data.evaluate(mol_only=False, match_type="hybrid")
        if hard_res.get("ignored"): continue
        fp_fn = res["fp"] + res["fn"]
        if fp_fn == 0: continue
        if vis_range and not (vis_range[0] <= fp_fn <= vis_range[1]): continue
        results.append({"idx": idx, "fn": fn, "data": data, "res": res,
                        "hard": hard_res, "soft": soft_res, "hybrid": hybrid_res})
    results.sort(key=lambda r: r["res"]["fp"] + r["res"]["fn"], reverse=True)
    if limit_vis > 0: results = results[:limit_vis]

    md_path = os.path.join(report_dir, f"{match_type}_match_errors.md")
    vis_count = 0
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(f"# {match_type.capitalize()}-Match Error Report\n\n")
        f.write(f"Generated: {ts} | Cases: {len(results)}\n\n---\n\n")
        for r in results:
            idx, fn, data = r["idx"], r["fn"], r["data"]
            for label, ev in [("Hard", r["hard"]), ("Soft", r["soft"]), ("Hybrid", r["hybrid"])]:
                m = compute_metrics(ev["tp"], ev["gold_total"], ev["pred_total"])
                f.write(f"**{label}**: TP={ev['tp']} FP={ev['fp']} FN={ev['fn']} "
                        f"P={m['precision']:.3f} R={m['recall']:.3f} F1={m['f1']:.3f}  \n")
            f.write("\n")
            gt = [rx for rx in data.gold_reactions if not rx.is_empty(mol_only)]
            pr = [rx for rx in data.pred_reactions if not rx.is_empty(mol_only)]
            cur = r["res"]; matched_gt = {m[0] for m in cur["matches"]}; matched_pr = {m[1] for m in cur["matches"]}
            pairs = []
            for gi, pi in cur["matches"]:
                pairs.append({"gt": gt[gi], "pred": pr[pi], "status": "Matched (TP)"})
            for pi in range(len(pr)):
                if pi not in matched_pr: pairs.append({"gt": None, "pred": pr[pi], "status": "Wrong Pred (FP)"})
            for gi in range(len(gt)):
                if gi not in matched_gt: pairs.append({"gt": gt[gi], "pred": None, "status": "Missed GT (FN)"})
            img = cv2.imread(data.image_path)
            if img is None: f.write(f"(Cannot read image: {data.image_path})\n\n"); continue
            img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            for k, pair in enumerate(pairs):
                vp = visualize_reaction_pair(img_rgb, pair, report_dir, fn, idx, k, mol_only, match_type)
                if vp:
                    rel = os.path.relpath(vp, report_dir).replace("\\", "/")
                    f.write(f"**{pair['status']}**\n\n![vis]({rel})\n\n")
                    vis_count += 1
            f.write("---\n\n")
    print(f"Visualization report: {md_path} ({vis_count} images)")

# --- Excel export ---
def save_match_results_to_excel(gold_images, pred_images, image_base_path, output_dir, filter_recover=False):
    if not HAS_XLSX:
        print("[WARN] openpyxl not installed. Skipping Excel export."); return
    print("\n" + "=" * 30 + " Exporting to Excel " + "=" * 30)
    modes = {"Hard": {"mol_only": False, "match_type": "hard"},
             "Soft": {"mol_only": True, "match_type": "soft"},
             "Hybrid": {"mol_only": False, "match_type": "hybrid"}}
    groups = ["Overall", "single", "multiple", "tree", "graph"]
    wb = openpyxl.Workbook()
    gbn = {g["file_name"]: g for g in gold_images}
    pbn = {p["file_name"]: p for p in pred_images}
    for k, (sn, params) in enumerate(modes.items()):
        ws = wb.active if k == 0 else wb.create_sheet(sn); ws.title = sn
        headers = ["", "TP", "FP", "FN", "Gold", "Pred", "P%", "R%", "F1%"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor="FFD700")
        for r, grp in enumerate(groups, 2):
            ws.cell(row=r, column=1, value=grp).font = Font(bold=True)
        all_res, group_res = [], defaultdict(list)
        for fn in sorted(set(gbn) & set(pbn)):
            data = ReactionImageData(gbn[fn], pbn[fn], image_base_path, filter_recover)
            res = data.evaluate(**params)
            if res.get("ignored"): continue
            all_res.append(res); grp = gbn[fn].get("diagram_type", "unknown"); group_res[grp].append(res)
        def wr(row, rl):
            if not rl: return
            tp = sum(r["tp"] for r in rl); gt = sum(r["gold_total"] for r in rl)
            pd_ = sum(r["pred_total"] for r in rl); fp = sum(r["fp"] for r in rl); fn_ = sum(r["fn"] for r in rl)
            m = compute_metrics(tp, gt, pd_)
            for c, v in enumerate([tp, fp, fn_, gt, pd_, round(m["precision"]*100,2), round(m["recall"]*100,2), round(m["f1"]*100,2)], 2):
                ws.cell(row=row, column=c, value=v)
        wr(2, all_res)
        for grp in ["single", "multiple", "tree", "graph"]:
            wr(groups.index(grp) + 2, group_res.get(grp, []))
        for col in range(1, 10): ws.column_dimensions[get_column_letter(col)].width = 15
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(output_dir, f"match_results_{ts}.xlsx"); wb.save(out_path)
    print(f"Saved: {out_path}\n" + "=" * 80)

# --- Per-image F1 CSV ---
def save_per_image_f1_csv(gold_images, pred_images, image_base_path, output_dir, filter_recover=False):
    gbn = {g["file_name"]: g for g in gold_images}
    pbn = {p["file_name"]: p for p in pred_images}
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    for sheet, params in [("soft", {"mol_only": True, "match_type": "soft"}),
                          ("hybrid", {"mol_only": False, "match_type": "hybrid"})]:
        rows = []
        for fn in sorted(set(gbn) & set(pbn)):
            data = ReactionImageData(gbn[fn], pbn[fn], image_base_path, filter_recover)
            res = data.evaluate(**params)
            if res.get("ignored"): continue
            m = compute_metrics(res["tp"], res["gold_total"], res["pred_total"])
            rows.append((fn, res["gold_total"], m["f1"]))
        rows.sort(key=lambda x: (x[2], x[0]))
        csv_path = os.path.join(output_dir, f"f1_ranking_{sheet}_{ts}.csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f); w.writerow(["File Name", "GT Count", "F1"])
            for fn, cnt, f1 in rows: w.writerow([fn, cnt, f"{f1:.6f}"])
        print(f"Saved: {csv_path}")

# --- CLI ---
def main():
    parser = argparse.ArgumentParser(description="Evaluate reaction extraction predictions.")
    parser.add_argument("--ground_truth_file", required=True)
    parser.add_argument("--pred_file", required=True)
    parser.add_argument("--image_base_path", required=True)
    parser.add_argument("--output_dir", default="./eval_results")
    parser.add_argument("--filter_recover", action="store_true")
    parser.add_argument("--limit_vis", type=int, default=-1, help="Max visualizations (<=0 = no limit)")
    parser.add_argument("--vis_range", type=int, nargs=2, metavar=("MIN","MAX"), default=None,
                        help="Only visualize cases where MIN <= (FP+FN) <= MAX")
    parser.add_argument("--mode", default="overall",
                        choices=["overall", "export_excel", "export_f1_rank", "visualize", "all"])
    args = parser.parse_args()
    os.makedirs(args.output_dir, exist_ok=True)
    log_file = setup_logging(args.output_dir)
    logging.info(f"Log: {log_file}")
    with open(args.ground_truth_file, encoding="utf-8") as f: gold = json.load(f)
    with open(args.pred_file, encoding="utf-8") as f: pred = json.load(f)
    gi, pi = gold.get("images", []), pred.get("images", [])
    if not gi: logging.error("No GT images."); return
    if not pi: logging.error("No pred images."); return
    common = {i["file_name"] for i in gi} & {i["file_name"] for i in pi}
    logging.info(f"GT: {len(gi)}, Pred: {len(pi)}, Common: {len(common)}")
    if not common: logging.error("No matching filenames."); return

    if args.mode in ("overall", "all"):
        run_overall_evaluation(gi, pi, args.image_base_path, args.filter_recover)
    if args.mode in ("export_excel", "all"):
        save_match_results_to_excel(gi, pi, args.image_base_path, args.output_dir, args.filter_recover)
    if args.mode in ("export_f1_rank", "all"):
        save_per_image_f1_csv(gi, pi, args.image_base_path, args.output_dir, args.filter_recover)
    if args.mode in ("visualize", "all"):
        for mt, mo in [("soft", True), ("hybrid", False)]:
            generate_visualization_report(gi, pi, args.image_base_path, args.output_dir,
                                           vis_range=args.vis_range, limit_vis=args.limit_vis,
                                           mol_only=mo, match_type=mt, filter_recover=args.filter_recover)

if __name__ == "__main__":
    main()
